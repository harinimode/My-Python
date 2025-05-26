import os
import cv2
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def preprocess_image(path, size=(128, 128)):
    img = cv2.imread(path, cv2.IMREAD_GRAYSCALE)
    if img is None:
        return None
    img = cv2.resize(img, size)
    return img.flatten() / 255.0

# Your dataset root folder
base_path = r"C:\Users\HARSHITHA\Downloads\Python\quality\neew"

results = []

for category in os.listdir(base_path):
    c_path = os.path.join(base_path, category)
    train_good = os.path.join(c_path, "train", "good")
    test_root  = os.path.join(c_path, "test")

    if not (os.path.isdir(train_good) and os.path.isdir(test_root)):
        continue

    for defect_type in os.listdir(test_root):
        d_dir = os.path.join(test_root, defect_type)
        for fname in os.listdir(d_dir):
            fpath = os.path.join(d_dir, fname)
            vec = preprocess_image(fpath)
            if vec is None:
                continue

            status = "approved" if defect_type.lower() == "good" else "rejected"

            results.append({
                "category": category,
                "defect_type": defect_type,
                "file_name": fname,
                "status": status
            })

# Save to Excel with colors
df = pd.DataFrame(results)
xls = "e3_final_one.xlsx"
df.to_excel(xls, index=False)

wb = load_workbook(xls)
ws = wb.active
green = PatternFill("solid", start_color="C6EFCE")
red   = PatternFill("solid", start_color="FFC7CE")

for cell in ws[1]:
    if cell.value == "status":
        status_col = cell.column_letter
        break

for row in range(2, ws.max_row + 1):
    c = ws[f"{status_col}{row}"]
    c.fill = green if c.value == "approved" else red

wb.save(xls)
print(f"✅ Done! All 'good' items approved, others rejected. Results in {xls}")
